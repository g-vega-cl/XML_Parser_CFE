import subprocess
import sys
"""
def pip_install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])


pip_install("pandas")
pip_install("selenium")
pip_install("webdriver_manager")
pip_install("openpyxl")
"""

import os
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Fill
import pandas as pd
import sys




if __name__ == "__main__":
	JS_value = sys.argv[1]
	#print(JS_value)

	folderString = JS_value[0] + ":\\" + JS_value[3:]

	wb = openpyxl.Workbook()

	sheet_obj = wb.active

	sheet_obj.title = "Recibos"

	greenFill = PatternFill(start_color='1fae00',
                   end_color='1fae00',
                   fill_type='solid')

	orangeFill = PatternFill(start_color='ff7931',
                   end_color='ff7931',
                   fill_type='solid')	

	fontBold = Font(name='Calibri',
	                 size=11,
	                 bold=True,
	                 italic=False,
	                 vertAlign=None,
	                 underline='none',
	                 strike=False,
	                 color='fbfbfb')        

	sheet_obj.column_dimensions['B'].width = 23
	sheet_obj.column_dimensions['C'].width = 23	                  	
	sheet_obj.column_dimensions['D'].width = 23
	sheet_obj.column_dimensions['E'].width = 23
	sheet_obj.column_dimensions['F'].width = 23
	sheet_obj.column_dimensions['G'].width = 23
	sheet_obj.column_dimensions['h'].width = 15	
	sheet_obj.sheet_view.showGridLines = False

	data_2017 = 0
	data_2018 = 0
	data_2019 = 0
	data_2020 = 0

	year_count = 0

	#os.chdir("XML_Aqui")
	os.chdir(folderString)

	for file in os.listdir():
		if(file == "Recibos.xlsx"):
			continue
		
		tree = ET.parse(file)
		root = tree.getroot()
		try:
			Datos = root[5][2]
		  # A veces el formato de los XML cambia...
		except:
			try:
				Datos = root[4][2]
			except:
				Datos = root[6][2]
		
		Fecha = Datos.findall("FECHASTA")[0].text	
		
		current_y_str = Fecha[-2:]
		if(current_y_str == "17"):
			if(data_2017 == 0):
				data_2017 = 1

		elif(current_y_str == "18"):
			if(data_2018 == 0):
				data_2018 = 1 + data_2017			

		elif(current_y_str == "19"):
			if(data_2019 == 0):
				data_2019 = 1 + data_2018
		
		elif(current_y_str == "20"):
			if(data_2020 == 0):
				data_2020 = 1 + data_2019			


	for file in os.listdir():

		if(file == "Recibos.xlsx"):
			continue

		tree = ET.parse(file)
		root = tree.getroot()

		try:
			Datos = root[5][2]
		  # A veces el formato de los XML cambia...
		except:
			try:
				Datos = root[4][2]
			except:
				Datos = root[6][2]

		Fecha = Datos.findall("FECHASTA")[0].text

		Punta = Datos.findall("CONSUMO1F")[0].text
		Intermedia = Datos.findall("CONSUMO2F")[0].text
		Base = Datos.findall("CONSUMO3F")[0].text

		Punta_c = Datos.findall("DEMANDA1P")[0].text
		Intermedia_c = Datos.findall("DEMANDA2P")[0].text
		Base_c = Datos.findall("DEMANDA3P")[0].text  

		total = Datos.findall("TOTAL_CENT_XML")[0].text

		if("ENE" in Fecha):
			Mes = "ENE"
			Num_celda = 4

		elif("FEB" in Fecha):
			Mes = "FEB"
			Num_celda = 5

		elif("MAR" in Fecha):
			Mes = "MAR"
			Num_celda = 6

		elif("ABR" in Fecha):
			Mes = "ABR"
			Num_celda = 7

		elif("MAY" in Fecha):
			Mes = "MAY"
			Num_celda = 8

		elif("JUN" in Fecha):
			Mes = "JUN"
			Num_celda = 9

		elif("JUL" in Fecha):
			Mes = "JUL"
			Num_celda = 10

		elif("AGO" in Fecha):
			Mes = "AGO"
			Num_celda = 11

		elif("SEP" in Fecha):
			Mes = "SEP"
			Num_celda = 12

		elif("OCT" in Fecha):
			Mes = "OCT"
			Num_celda = 13

		elif("NOV" in Fecha):
			Mes = "NOV"
			Num_celda = 14

		elif("DIC" in Fecha):
			Mes = "DIC"
			Num_celda = 15


		for i in range(8):
			if(Fecha[-2:] == "17"):
				year_str = "2017"
				if(data_2017 > 0):
					Celda_Head = 4 + 16 * (data_2017 - 1)

			elif(Fecha[-2:] == "18"):
				Num_celda_curr = Num_celda + 16 * (data_2018 - 1)
				Celda_Head = 4 + 16 * (data_2018 - 1)
				year_str = "2018"

			elif(Fecha[-2:] == "19"):
				Num_celda_curr = Num_celda + 16 * (data_2019 - 1)
				Celda_Head = 4 + 16 * (data_2019 - 1)
				year_str = "2019"

			elif(Fecha[-2:] == "20"):
				Num_celda_curr = Num_celda + 16 * (data_2020 - 1)
				Celda_Head = 4 + 16 * (data_2020 - 1)
				year_str = "2020"

			sheet_obj.merge_cells('A{}:H{}'.format(str(Celda_Head-2),str(Celda_Head-2)))
			sheet_obj["A{}".format(str(Celda_Head-2))] = "Recibos {}".format(year_str)
			sheet_obj["A{}".format(str(Celda_Head-2))].font = fontBold
			sheet_obj["A{}".format(str(Celda_Head-2))].fill = greenFill

			sheet_obj["B{}".format(str(Celda_Head-1))] = "Capacidad Base"
			sheet_obj["C{}".format(str(Celda_Head-1))] = "Capacidad Intermedia"	
			sheet_obj["D{}".format(str(Celda_Head-1))] = "Capacidad Punta"		
			
			sheet_obj["E{}".format(str(Celda_Head-1))] = "Energia Base"	
			sheet_obj["F{}".format(str(Celda_Head-1))] = "Energia Intermedia"	
			sheet_obj["G{}".format(str(Celda_Head-1))] = "Energia Punta"	
			sheet_obj["H{}".format(str(Celda_Head-1))] = "Total"	

			if(i == 0):
				columna = "A"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = Mes
			if(i == 1):
				columna = "B"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = float(Base_c)				

			if(i == 2):
				columna = "C"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = float(Intermedia_c)

			if(i == 3):
				columna = "D"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = float(Punta_c)

			if(i == 4):
				columna = "E"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = float(Base)		

			if(i == 5):
				columna = "F"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = float(Intermedia)			

			if(i == 6):
				columna = "G"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				sheet_obj[Celda_a_cambiar] = float(Punta)		

			if(i == 7):
				columna = "H"
				Celda_a_cambiar = "{}{}".format(columna,str(Num_celda_curr))
				total_int = float(total.replace(',',''))
				sheet_obj[Celda_a_cambiar] = total_int	

	#os.chdir("../") 
	wb.save("Recibos.xlsx")


